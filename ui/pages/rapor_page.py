"""
VD3350 Manager - Raporlama Sayfası
=====================================
Günlük/haftalık/aylık üretim, fire ve performans raporları.
PDF ve Excel dışa aktarma.
"""

from typing import Optional
from datetime import datetime, timedelta
from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QPushButton, QComboBox, QFrame, QTableWidget,
    QTableWidgetItem, QHeaderView, QAbstractItemView,
    QDateEdit, QMessageBox, QFileDialog, QScrollArea
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QFont

from ui.widgets.common import SectionHeader, Separator
from services.is_emri_service import IsEmriService
from services.fire_service import FireService
from database import db


class RaporPage(QWidget):
    """Raporlama sayfası."""

    def __init__(self, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        self.is_svc = IsEmriService()
        self.fire_svc = FireService()
        self._build_ui()
        self._generate_report()

    def _build_ui(self) -> None:
        """Sayfa arayüzü."""
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(32, 24, 32, 24)
        main_layout.setSpacing(24)

        # Başlık
        header_row = QHBoxLayout()
        header = SectionHeader("📊 Raporlama", "Üretim ve fire raporları")
        header_row.addWidget(header)
        main_layout.addLayout(header_row)
        main_layout.addWidget(Separator())

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)

        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setSpacing(20)
        content_layout.setContentsMargins(0, 0, 0, 0)

        # Filtre kartı
        content_layout.addWidget(self._build_filter_card())

        # Özet istatistik kartları
        content_layout.addWidget(self._build_summary_card())

        # Top listeler
        tops_row = QHBoxLayout()
        tops_row.setSpacing(16)
        tops_row.addWidget(self._build_top_musteri_card())
        tops_row.addWidget(self._build_top_malzeme_card())
        tops_row.addWidget(self._build_top_neden_card())
        content_layout.addLayout(tops_row)

        # Detaylı iş emri tablosu
        content_layout.addWidget(QLabel("📋 İş Emri Detayları"))
        content_layout.addWidget(self._build_detail_table())

        scroll.setWidget(content)
        main_layout.addWidget(scroll)

    def _build_filter_card(self) -> QFrame:
        """Filtre ve dışa aktarma kartı."""
        card = QFrame()
        card.setObjectName("card")
        layout = QHBoxLayout(card)
        layout.setContentsMargins(20, 14, 20, 14)
        layout.setSpacing(16)

        # Dönem seçimi
        period_lbl = QLabel("Dönem:")
        period_lbl.setFont(QFont("Segoe UI", 12))
        layout.addWidget(period_lbl)

        self.period_combo = QComboBox()
        self.period_combo.addItems([
            "Bugün", "Bu Hafta", "Bu Ay", "Son 30 Gün",
            "Son 90 Gün", "Tüm Zamanlar"
        ])
        self.period_combo.currentTextChanged.connect(self._on_period_changed)
        layout.addWidget(self.period_combo)

        # Özel tarih aralığı
        self.baslangic_date = QDateEdit()
        self.baslangic_date.setDate(QDate.currentDate().addDays(-30))
        self.baslangic_date.setCalendarPopup(True)
        self.baslangic_date.setDisplayFormat("dd.MM.yyyy")
        self.baslangic_date.setEnabled(False)
        layout.addWidget(QLabel("—"))
        layout.addWidget(self.baslangic_date)

        self.bitis_date = QDateEdit()
        self.bitis_date.setDate(QDate.currentDate())
        self.bitis_date.setCalendarPopup(True)
        self.bitis_date.setDisplayFormat("dd.MM.yyyy")
        self.bitis_date.setEnabled(False)
        layout.addWidget(self.bitis_date)

        # Rapor oluştur
        self.generate_btn = QPushButton("🔄 Raporu Güncelle")
        self.generate_btn.setObjectName("primaryBtn")
        self.generate_btn.clicked.connect(self._generate_report)
        layout.addWidget(self.generate_btn)

        layout.addStretch()

        # Dışa aktarma
        self.pdf_btn = QPushButton("📄 PDF İndir")
        self.pdf_btn.setObjectName("secondaryBtn")
        self.pdf_btn.clicked.connect(self._export_pdf)
        layout.addWidget(self.pdf_btn)

        self.excel_btn = QPushButton("📊 Excel İndir")
        self.excel_btn.setObjectName("secondaryBtn")
        self.excel_btn.clicked.connect(self._export_excel)
        layout.addWidget(self.excel_btn)

        return card

    def _build_summary_card(self) -> QFrame:
        """Özet istatistik kartı."""
        card = QFrame()
        card.setObjectName("card")
        layout = QGridLayout(card)
        layout.setContentsMargins(20, 16, 20, 16)
        layout.setSpacing(16)

        title = QLabel("📈 Özet İstatistikler")
        title.setFont(QFont("Segoe UI", 13, QFont.Weight.Bold))
        layout.addWidget(title, 0, 0, 1, 6)

        def stat_widget(label: str, color: str) -> tuple[QLabel, QLabel]:
            lbl_title = QLabel(label)
            lbl_title.setObjectName("labelMuted")
            lbl_title.setFont(QFont("Segoe UI", 10))

            lbl_val = QLabel("—")
            lbl_val.setFont(QFont("Segoe UI", 20, QFont.Weight.ExtraBold))
            lbl_val.setStyleSheet(
                f"color: {color}; background: transparent; border: none;"
            )
            return lbl_title, lbl_val

        stats = [
            ("Toplam İş", "#3b82f6"),
            ("Toplam Metraj", "#10b981"),
            ("Tamamlanan", "#10b981"),
            ("Hatalı Kesim", "#ef4444"),
            ("Toplam Fire (m)", "#f97316"),
            ("Fire Oranı (%)", "#ef4444"),
        ]

        self._stat_labels: dict[str, QLabel] = {}
        for col, (label, color) in enumerate(stats):
            t, v = stat_widget(label, color)
            layout.addWidget(t, 1, col)
            layout.addWidget(v, 2, col)
            self._stat_labels[label] = v

        return card

    def _build_top_musteri_card(self) -> QFrame:
        """En çok kesilen müşteriler kartı."""
        card = QFrame()
        card.setObjectName("card")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(8)

        title = QLabel("🏆 En Çok Kesilen Müşteriler")
        title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(title)

        self.top_musteri_layout = QVBoxLayout()
        layout.addLayout(self.top_musteri_layout)

        return card

    def _build_top_malzeme_card(self) -> QFrame:
        """En çok kullanılan malzemeler kartı."""
        card = QFrame()
        card.setObjectName("card")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(8)

        title = QLabel("📦 En Çok Kullanılan Malzeme")
        title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(title)

        self.top_malzeme_layout = QVBoxLayout()
        layout.addLayout(self.top_malzeme_layout)

        return card

    def _build_top_neden_card(self) -> QFrame:
        """En çok görülen fire nedenleri kartı."""
        card = QFrame()
        card.setObjectName("card")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(16, 14, 16, 14)
        layout.setSpacing(8)

        title = QLabel("⚠️ En Çok Görülen Fire Nedeni")
        title.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        layout.addWidget(title)

        self.top_neden_layout = QVBoxLayout()
        layout.addLayout(self.top_neden_layout)

        return card

    def _build_detail_table(self) -> QTableWidget:
        """Detaylı iş emri tablosu."""
        headers = [
            "Form No", "Tarih", "Müşteri", "Malzeme",
            "Metraj (m)", "Adet", "Durum"
        ]
        self.detail_table = QTableWidget()
        self.detail_table.setColumnCount(len(headers))
        self.detail_table.setHorizontalHeaderLabels(headers)
        self.detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.detail_table.setAlternatingRowColors(True)
        self.detail_table.verticalHeader().setVisible(False)
        self.detail_table.setShowGrid(False)
        self.detail_table.setMaximumHeight(350)

        header = self.detail_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)

        return self.detail_table

    def _get_date_range(self) -> tuple[str, str]:
        """Seçilen döneme göre tarih aralığı döner."""
        bugun = datetime.now().date()
        period = self.period_combo.currentText()

        if period == "Bugün":
            return str(bugun), str(bugun)
        elif period == "Bu Hafta":
            pazartesi = bugun - timedelta(days=bugun.weekday())
            return str(pazartesi), str(bugun)
        elif period == "Bu Ay":
            ay_basi = bugun.replace(day=1)
            return str(ay_basi), str(bugun)
        elif period == "Son 30 Gün":
            return str(bugun - timedelta(days=30)), str(bugun)
        elif period == "Son 90 Gün":
            return str(bugun - timedelta(days=90)), str(bugun)
        elif period == "Özel Aralık":
            return (
                self.baslangic_date.date().toString("yyyy-MM-dd"),
                self.bitis_date.date().toString("yyyy-MM-dd"),
            )
        else:  # Tüm Zamanlar
            return "2000-01-01", str(bugun)

    def _on_period_changed(self, period: str) -> None:
        """Dönem değişince tarih alanlarını aktif/pasif yapar."""
        ozel = period == "Özel Aralık"
        self.baslangic_date.setEnabled(ozel)
        self.bitis_date.setEnabled(ozel)

    def _generate_report(self) -> None:
        """Raporu oluşturur ve günceller."""
        baslangic, bitis = self._get_date_range()

        # İş emirleri
        rows = db.fetchall(
            """SELECT * FROM is_emirleri
               WHERE tarih BETWEEN ? AND ?
               ORDER BY tarih DESC""",
            (baslangic, bitis),
        )

        toplam_is = len(rows)
        toplam_metraj = sum(r["metraj"] for r in rows)
        tamamlanan = sum(1 for r in rows if r["durum"] == "Tamamlandı")
        hatali = sum(1 for r in rows if r["durum"] == "Hatalı Kesim")

        # Fire
        fire_rows = db.fetchall(
            """SELECT * FROM fire_ve_hatalar
               WHERE tarih BETWEEN ? AND ?""",
            (baslangic, bitis),
        )
        toplam_fire = sum(r["hatali_metre"] or 0 for r in fire_rows)
        fire_orani = (toplam_fire / toplam_metraj * 100) if toplam_metraj > 0 else 0.0

        # Stat güncelle
        vals = {
            "Toplam İş": str(toplam_is),
            "Toplam Metraj": f"{toplam_metraj:,.0f} m",
            "Tamamlanan": str(tamamlanan),
            "Hatalı Kesim": str(hatali),
            "Toplam Fire (m)": f"{toplam_fire:,.1f}",
            "Fire Oranı (%)": f"%{fire_orani:.1f}",
        }
        for key, val in vals.items():
            if key in self._stat_labels:
                self._stat_labels[key].setText(val)

        # Top müşteriler
        self._update_top_list(
            db.fetchall(
                """SELECT musteri_adi, SUM(metraj) as toplam, COUNT(*) as adet
                   FROM is_emirleri WHERE tarih BETWEEN ? AND ?
                   GROUP BY musteri_adi ORDER BY toplam DESC LIMIT 5""",
                (baslangic, bitis),
            ),
            self.top_musteri_layout,
            "musteri_adi",
            "toplam",
            "#3b82f6",
        )

        # Top malzemeler
        self._update_top_list(
            db.fetchall(
                """SELECT malzeme_cinsi, SUM(metraj) as toplam, COUNT(*) as adet
                   FROM is_emirleri WHERE tarih BETWEEN ? AND ?
                   GROUP BY malzeme_cinsi ORDER BY toplam DESC LIMIT 5""",
                (baslangic, bitis),
            ),
            self.top_malzeme_layout,
            "malzeme_cinsi",
            "toplam",
            "#10b981",
        )

        # Top fire nedenleri
        self._update_top_list(
            db.fetchall(
                """SELECT fire_nedeni, COUNT(*) as adet, SUM(hatali_metre) as toplam
                   FROM fire_ve_hatalar WHERE tarih BETWEEN ? AND ?
                   AND fire_nedeni IS NOT NULL
                   GROUP BY fire_nedeni ORDER BY adet DESC LIMIT 5""",
                (baslangic, bitis),
            ),
            self.top_neden_layout,
            "fire_nedeni",
            "adet",
            "#f59e0b",
            suffix=" kez",
        )

        # Detay tablosu
        self.detail_table.setRowCount(len(rows))
        durum_colors = {
            "Bekliyor": "#f59e0b",
            "Devam Ediyor": "#3b82f6",
            "Tamamlandı": "#10b981",
            "Hatalı Kesim": "#ef4444",
        }
        for row_idx, r in enumerate(rows):
            cells = [
                r["form_numarasi"],
                r["tarih"],
                r["musteri_adi"],
                r["malzeme_cinsi"],
                f"{r['metraj']:,.1f}",
                f"{r['adet']:,}",
                r["durum"],
            ]
            for col, text in enumerate(cells):
                item = QTableWidgetItem(text)
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if col == 6:
                    from PyQt6.QtGui import QColor
                    item.setForeground(QColor(durum_colors.get(text, "#94a3b8")))
                    item.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
                self.detail_table.setItem(row_idx, col, item)
            self.detail_table.setRowHeight(row_idx, 36)

        # Rapor verilerini sakla (export için)
        self._report_rows = [dict(r) for r in rows]
        self._report_stats = {
            "baslangic": baslangic,
            "bitis": bitis,
            "toplam_is": toplam_is,
            "toplam_metraj": toplam_metraj,
            "tamamlanan": tamamlanan,
            "hatali": hatali,
            "toplam_fire": toplam_fire,
            "fire_orani": fire_orani,
        }

    def _update_top_list(
        self, rows, layout: QVBoxLayout, name_col: str,
        val_col: str, color: str, suffix: str = " m"
    ) -> None:
        """Top liste widgetlarını günceller."""
        # Temizle
        while layout.count():
            item = layout.takeAt(0)
            if item and item.widget():
                item.widget().deleteLater()

        if not rows:
            lbl = QLabel("Veri yok")
            lbl.setObjectName("labelMuted")
            layout.addWidget(lbl)
            return

        max_val = max(r[val_col] or 0 for r in rows) or 1

        for i, r in enumerate(rows):
            val = r[val_col] or 0
            pct = int((val / max_val) * 100)

            row_widget = QWidget()
            row_layout = QVBoxLayout(row_widget)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(2)

            name_val = QHBoxLayout()
            name_lbl = QLabel(f"{i+1}. {r[name_col]}")
            name_lbl.setFont(QFont("Segoe UI", 11))
            val_lbl = QLabel(f"{val:,.0f}{suffix}")
            val_lbl.setFont(QFont("Segoe UI", 11, QFont.Weight.Bold))
            val_lbl.setStyleSheet(
                f"color: {color}; background: transparent; border: none;"
            )
            name_val.addWidget(name_lbl)
            name_val.addStretch()
            name_val.addWidget(val_lbl)
            row_layout.addLayout(name_val)

            from PyQt6.QtWidgets import QProgressBar
            pb = QProgressBar()
            pb.setRange(0, 100)
            pb.setValue(pct)
            pb.setTextVisible(False)
            pb.setFixedHeight(4)
            pb.setStyleSheet(
                f"QProgressBar {{ background: #334155; border-radius: 2px; border: none; }}"
                f"QProgressBar::chunk {{ background: {color}; border-radius: 2px; }}"
            )
            row_layout.addWidget(pb)
            layout.addWidget(row_widget)

    def _export_excel(self) -> None:
        """Excel'e dışa aktarır."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.warning(
                self, "Hata",
                "Excel dışa aktarma için 'openpyxl' kütüphanesi gerekli.\n"
                "pip install openpyxl"
            )
            return

        dosya, _ = QFileDialog.getSaveFileName(
            self, "Excel Kaydet", "VD3350_Rapor.xlsx",
            "Excel (*.xlsx)"
        )
        if not dosya:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Üretim Raporu"  # type: ignore

        # Başlık
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_alignment = Alignment(horizontal="center", vertical="center")

        headers = ["Form No", "Tarih", "Müşteri", "Malzeme", "Metraj (m)", "Adet", "Durum"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)  # type: ignore
            cell.font = header_font  # type: ignore
            cell.fill = header_fill  # type: ignore
            cell.alignment = header_alignment  # type: ignore

        # Veriler
        for row_idx, r in enumerate(self._report_rows, 2):
            ws.cell(row=row_idx, column=1, value=r["form_numarasi"])  # type: ignore
            ws.cell(row=row_idx, column=2, value=r["tarih"])  # type: ignore
            ws.cell(row=row_idx, column=3, value=r["musteri_adi"])  # type: ignore
            ws.cell(row=row_idx, column=4, value=r["malzeme_cinsi"])  # type: ignore
            ws.cell(row=row_idx, column=5, value=r["metraj"])  # type: ignore
            ws.cell(row=row_idx, column=6, value=r["adet"])  # type: ignore
            ws.cell(row=row_idx, column=7, value=r["durum"])  # type: ignore

        # Sütun genişlikleri
        for col in ws.columns:  # type: ignore
            max_length = max(len(str(cell.value or "")) for cell in col)  # type: ignore
            ws.column_dimensions[col[0].column_letter].width = max(12, max_length + 4)  # type: ignore

        wb.save(dosya)
        QMessageBox.information(self, "Başarılı", f"Excel kaydedildi:\n{dosya}")

    def _export_pdf(self) -> None:
        """PDF'e dışa aktarır."""
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            QMessageBox.warning(
                self, "Hata",
                "PDF dışa aktarma için 'reportlab' kütüphanesi gerekli.\n"
                "pip install reportlab"
            )
            return

        dosya, _ = QFileDialog.getSaveFileName(
            self, "PDF Kaydet", "VD3350_Rapor.pdf",
            "PDF (*.pdf)"
        )
        if not dosya:
            return

        doc = SimpleDocTemplate(dosya, pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story = []

        # Başlık
        story.append(Paragraph("VD3350 Üretim Raporu — Gündoğdu Kağıt", styles["Title"]))
        story.append(
            Paragraph(
                f"Tarih Aralığı: {self._report_stats['baslangic']} — {self._report_stats['bitis']}",
                styles["Normal"]
            )
        )
        story.append(Spacer(1, 12))

        # Özet
        ozet = [
            ["Toplam İş", "Toplam Metraj", "Tamamlanan", "Hatalı", "Toplam Fire", "Fire Oranı"],
            [
                str(self._report_stats["toplam_is"]),
                f"{self._report_stats['toplam_metraj']:,.0f} m",
                str(self._report_stats["tamamlanan"]),
                str(self._report_stats["hatali"]),
                f"{self._report_stats['toplam_fire']:,.1f} m",
                f"%{self._report_stats['fire_orani']:.1f}",
            ]
        ]
        ozet_tablo = Table(ozet, colWidths=[100] * 6)
        ozet_tablo.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(ozet_tablo)
        story.append(Spacer(1, 16))

        # Detay tablosu
        story.append(Paragraph("İş Emri Detayları", styles["Heading2"]))
        story.append(Spacer(1, 6))

        headers = ["Form No", "Tarih", "Müşteri", "Malzeme", "Metraj (m)", "Adet", "Durum"]
        data = [headers]
        for r in self._report_rows:
            data.append([
                r["form_numarasi"],
                r["tarih"],
                r["musteri_adi"][:20],
                r["malzeme_cinsi"],
                f"{r['metraj']:,.1f}",
                f"{r['adet']:,}",
                r["durum"],
            ])

        col_widths = [80, 65, 130, 80, 70, 60, 90]
        tablo = Table(data, colWidths=col_widths)
        tablo.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e3a5f")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]))
        story.append(tablo)

        doc.build(story)
        QMessageBox.information(self, "Başarılı", f"PDF kaydedildi:\n{dosya}")

    def refresh(self) -> None:
        """Raporu yeniler."""
        self._generate_report()
